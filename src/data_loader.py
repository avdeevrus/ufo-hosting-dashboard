"""
Загрузка и нормализация данных UFO Hosting.

Источники:
- data/orders/*.csv  — выгрузки оплат (поштучно)
- data/ads/*.xlsx    — отчёты по рекламным кампаниям Яндекс.Директ (помесячно)
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

ORDERS_DIR = Path(__file__).resolve().parent.parent / "data" / "orders"
ADS_DIR = Path(__file__).resolve().parent.parent / "data" / "ads"

RU_MONTHS = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
    "май": 5, "июнь": 6, "июль": 7, "август": 8,
    "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}

EMAIL_RE = re.compile(r"\(([^)]+@[^)]+)\)")
MONEY_RE = re.compile(r"[\d\s.,]+")
BRACKET_RE = re.compile(r"\[([^\]]+)\]")


# ---------- ORDERS ----------

def _parse_money(value) -> float:
    """'20661.19 RUB' / '232 801,40' / float → float."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return np.nan
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return np.nan
    m = MONEY_RE.search(s)
    if not m:
        return np.nan
    num = m.group(0).strip().replace(" ", "").replace("\xa0", "")
    if "," in num and "." in num:
        num = num.replace(",", "")
    else:
        num = num.replace(",", ".")
    try:
        return float(num)
    except ValueError:
        return np.nan


def _extract_email(account: str) -> str | None:
    if not isinstance(account, str):
        return None
    m = EMAIL_RE.search(account)
    return m.group(1).strip().lower() if m else None


def _extract_name(account: str) -> str | None:
    if not isinstance(account, str):
        return None
    return EMAIL_RE.sub("", account).strip() or None


def _classify_product(name: str) -> tuple[str, str]:
    """Возвращает (семейство, локация)."""
    if not isinstance(name, str) or not name:
        return ("Прочее", "—")
    s = name.strip()
    bracket = BRACKET_RE.search(s)
    location = bracket.group(1).strip() if bracket else "—"
    base = BRACKET_RE.sub("", s).strip().rstrip("-").strip()

    low = base.lower()
    if "vpn" in low:
        family = "VPN"
    elif "hi-cpu" in low or "hi cpu" in low:
        family = "Hi-CPU VPS"
    elif "dedicated" in low or "выделен" in low:
        family = "Выделенный сервер"
    elif "hosting" in low or "хостинг" in low:
        family = "Хостинг"
    elif "domain" in low or "домен" in low:
        family = "Домен"
    elif "license" in low or "лиценз" in low or "ispmanager" in low:
        family = "Лицензии / аддоны"
    elif "commission" in low or "комисси" in low:
        family = "Комиссия платёжной системы"
    else:
        family = base.split()[0] if base else "Прочее"
    return (family, location)


def load_orders() -> pd.DataFrame:
    """Грузим все CSV из data/orders/, объединяем и дедуплицируем по ID покупки."""
    files = sorted(ORDERS_DIR.glob("*.csv"))
    frames: list[pd.DataFrame] = []
    for fp in files:
        df = pd.read_csv(fp, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        df.columns = [c.strip().strip('"').strip() for c in df.columns]
        if df.columns[0] != "ID покупки":
            df = df.rename(columns={df.columns[0]: "ID покупки"})
        df = df[df["ID покупки"].str.strip() != "Итого и средние"].copy()
        df["__source__"] = fp.name
        frames.append(df)
    if not frames:
        return pd.DataFrame()

    raw = pd.concat(frames, ignore_index=True)
    raw = raw.drop_duplicates(subset=["ID покупки"], keep="last")

    out = pd.DataFrame(
        {
            "order_id": raw["ID покупки"].astype(str),
            "product": raw["Товар в заказе"],
            "items_count": pd.to_numeric(raw["Товаров куплено"], errors="coerce"),
            "purchase_amount_rub": pd.to_numeric(raw["Сумма покупки, RUB"], errors="coerce"),
            "payment_status": raw["Статус оплаты"],
            "payment_date": pd.to_datetime(raw["Дата платежа"], errors="coerce"),
            "payment_amount": raw["Сумма оплаты"].apply(_parse_money),
            "account_raw": raw["Аккаунт"],
            "registration_date": pd.to_datetime(raw["Дата регистрации"], errors="coerce"),
            "new_or_old": raw["Новый/Старый"],
            "services_count": pd.to_numeric(raw["Кол-во услуг"], errors="coerce"),
            "services_detail": raw.get("Услуги (expense_locale_name)", ""),
        }
    )

    out["email"] = out["account_raw"].apply(_extract_email)
    out["client_name"] = out["account_raw"].apply(_extract_name)
    out["client_key"] = out["email"].fillna(out["account_raw"])
    fam_loc = out["product"].apply(_classify_product)
    out["product_family"] = [x[0] for x in fam_loc]
    out["product_location"] = [x[1] for x in fam_loc]
    out["is_paid"] = out["payment_status"].str.strip().str.lower().eq("оплачен")
    out["payment_month"] = out["payment_date"].dt.to_period("M").dt.to_timestamp()
    out["registration_month"] = out["registration_date"].dt.to_period("M").dt.to_timestamp()
    return out


# ---------- ADS ----------

def _parse_sheet_month(sheet: str) -> pd.Timestamp | None:
    """'Апрель 2026' → 2026-04-01."""
    s = sheet.strip().lower()
    parts = s.split()
    if len(parts) < 2:
        return None
    month = RU_MONTHS.get(parts[0])
    if not month:
        return None
    try:
        year = int(parts[1])
    except ValueError:
        return None
    return pd.Timestamp(year=year, month=month, day=1)


def _find_campaign_headers(ws) -> list[tuple[int, int, int]]:
    """Находим ВСЕ строки-заголовки таблиц кампаний на листе.
    Один лист может содержать несколько таблиц (напр. ноябрь: старый+новый аккаунт).
    Возвращаем список (row, name_col, spend_col)."""
    name_aliases = {"кампания", "название кампании"}
    spend_aliases = {"расход", "расход, руб.", "расход руб."}

    headers = []
    for r in range(1, ws.max_row + 1):
        cells = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        name_col = spend_col = None
        for i, v in enumerate(cells):
            if not isinstance(v, str):
                continue
            low = v.strip().lower()
            if low in name_aliases:
                name_col = i
            elif low in spend_aliases:
                spend_col = i
        if name_col is not None and spend_col is not None:
            headers.append((r, name_col, spend_col))
    return headers


def _read_campaigns_block(ws, header_row: int, name_col: int, spend_col: int,
                          stop_row: int | None = None) -> list[tuple[str, float]]:
    rows = []
    last_row = stop_row if stop_row is not None else ws.max_row
    for r in range(header_row + 1, last_row + 1):
        line = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        name = line[name_col] if name_col < len(line) else None
        spend = line[spend_col] if spend_col < len(line) else None
        if not isinstance(name, str) or not name.strip():
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in line):
                # пустая строка — окончание текущего блока
                break
            continue
        low = name.strip().lower()
        if low.startswith(("итого", "всего", "сводка")):
            break
        amount = _parse_money(spend)
        if np.isnan(amount):
            continue
        rows.append((name.strip(), amount))
    return rows


def load_ads() -> pd.DataFrame:
    """Грузим все XLSX из data/ads/, парсим каждый лист (месяц)."""
    files = sorted(ADS_DIR.glob("*.xlsx"))
    records = []
    for fp in files:
        wb = openpyxl.load_workbook(fp, data_only=True)
        for sheet_name in wb.sheetnames:
            month_dt = _parse_sheet_month(sheet_name)
            if month_dt is None:
                continue
            ws = wb[sheet_name]
            headers = _find_campaign_headers(ws)
            if not headers:
                continue
            for idx, hdr in enumerate(headers):
                stop = headers[idx + 1][0] - 1 if idx + 1 < len(headers) else None
                for campaign, spend in _read_campaigns_block(ws, *hdr, stop_row=stop):
                    records.append({
                        "month": month_dt,
                        "campaign": campaign,
                        "spend_rub": spend,
                        "source_file": fp.name,
                        "source_sheet": sheet_name,
                    })
        wb.close()
    return pd.DataFrame.from_records(records)


def ads_monthly_totals(ads: pd.DataFrame) -> pd.DataFrame:
    if ads.empty:
        return pd.DataFrame(columns=["month", "spend_rub", "campaigns"])
    g = ads.groupby("month", as_index=False).agg(
        spend_rub=("spend_rub", "sum"),
        campaigns=("campaign", "count"),
    )
    return g.sort_values("month")
